from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_system_owner_template():
    """Create System Owner Template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "System Owner Data"
    
    # Define headers
    headers = [
        "System Name",
        "Owner Name",
        "Owner Email",
        "Annual Licensing Cost",
        "Maintenance Cost",
        "Infrastructure Cost",
        "Internal Support Hours/Month",
        "External Support Cost"
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add sample rows for reference
    for row in range(2, 12):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # Set column widths
    column_widths = [20, 20, 25, 22, 18, 20, 28, 22]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save("System_Owner_Template.xlsx")
    print("✓ System_Owner_Template.xlsx created")

def create_department_usage_template():
    """Create Department Usage Template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Department Usage Data"
    
    # Define headers
    headers = [
        "Department Name",
        "System Used",
        "User Count",
        "Hidden Costs (Workarounds, Manual Effort)",
        "Estimated Effort (Hours/Month)",
        "Best Source for Unknowns"
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add sample rows
    for row in range(2, 12):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # Set column widths
    column_widths = [22, 20, 12, 35, 28, 25]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save("Department_Usage_Template.xlsx")
    print("✓ Department_Usage_Template.xlsx created")

def create_integration_owner_template():
    """Create Integration Owner Template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Integration Data"
    
    # Define headers
    headers = [
        "Interface (From → To)",
        "Method",
        "Frequency (Per Month)",
        "Cost",
        "Failure Incidents/Month",
        "Average MTTR (Hours)"
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add sample rows
    for row in range(2, 12):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # Set column widths
    column_widths = [30, 18, 22, 15, 25, 25]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save("Integration_Owner_Template.xlsx")
    print("✓ Integration_Owner_Template.xlsx created")

# Generate all three templates
if __name__ == "__main__":
    print("Generating Excel templates...\n")
    create_system_owner_template()
    create_department_usage_template()
    create_integration_owner_template()
    print("\nAll templates created successfully!")