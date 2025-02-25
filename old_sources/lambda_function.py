import aiohttp
import asyncio
import base64
import boto3
import concurrent.futures
import csv
import datetime
import json
import logging
import markdown2
import mimetypes
import nltk
import openai
import openpyxl
import os
import PyPDF2
import random
import re
import requests
import tempfile
import textwrap
import time
import wave
import zipfile

from aiohttp.client_exceptions import ClientConnectorSSLError
from boto3.dynamodb.conditions import Key
from botocore.exceptions import NoCredentialsError
from botocore.exceptions import ClientError
from bs4 import BeautifulSoup
from bson import ObjectId
from collections import Counter
from datetime import timedelta
from decimal import Decimal
from docx import Document
from fpdf import FPDF
from io import BytesIO, StringIO
from odoo_functions import authenticate, get_models, get_fields, create_record, fetch_records, update_record, delete_records, print_record
from openai import OpenAIError, BadRequestError
from prompts import prompts  # Import the prompts from prompts.py
from semantic_router import Route
from semantic_router import RouteLayer
from semantic_router.encoders import OpenAIEncoder
from tools import tools #Import tools from tools.py
from urllib.parse import urlparse, unquote, quote_plus, urlencode, urljoin
from xml.etree import ElementTree

from extservices import get_coordinates
from extservices import get_weather_data
from extservices import solve_math_problem 

from nltk.tokenize import sent_tokenize, word_tokenize

nltk.data.path.append("/opt/python/nltk_data") 

# Initialize API keys from environment variables              
slack_bot_token = os.getenv('SLACK_BOT_TOKEN')
google_api_key = os.getenv('GOOGLE_API_KEY')
gemini_api_key = os.getenv('GEMINI_API_KEY')
calendar_id = os.getenv('GOOGLE_CALENDAR_ID')      
erpnext_api_key = os.getenv('ERPNEXT_API_KEY') 
erpnext_api_secret = os.getenv('ERPNEXT_API_SECRET') 
openai_api_key = os.getenv('OPENAI_API_KEY')

odoo_url = "http://64.227.44.107:8069"
odoo_db = "Production"
odoo_login = "ai_bot"
odoo_password = "Carbon123#"

# Evomi proxy configuration
# Proxy URL
#proxy_url = f"http://core-residential.evomi.com:1000:aizukanne3:Ng8qM7DCChitRRuGDusL_country-US,CA"
proxy_url = "http://aizukanne3:Ng8qM7DCChitRRuGDusL_country-US,CA@core-residential.evomi.com:1000"


USER_AGENTS = [
    # Chrome (Windows, macOS, Linux)
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36",

    # Firefox (Windows, macOS, Linux)
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:93.0) Gecko/20100101 Firefox/93.0",
    "Mozilla/5.0 (X11; Linux x86_64; rv:95.0) Gecko/20100101 Firefox/95.0",

    # Safari (macOS, iOS)
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Version/14.1.2 Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1",

    # Edge (Windows)
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.64",

    # Opera
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 OPR/77.0.4054.277",

    # Mobile Browsers (Android, iOS)
    "Mozilla/5.0 (Linux; Android 11; SM-G998B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 10; Pixel 4 XL) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 14_4_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (iPad; CPU OS 14_7_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1 Mobile/15E148 Safari/604.1"
]


# Initialize OPENAI API keys  
client = openai.OpenAI(
    # defaults to os.environ.get("OPENAI_API_KEY")
    api_key = os.getenv('OPENAI_API_KEY')
)

encoder = OpenAIEncoder(
    os.environ.get("OPENAI_API_KEY")
)

# Initialize DynamoDB  
dynamodb = boto3.resource('dynamodb')

# Connect to your tables  
user_table = dynamodb.Table('staff_history')
assistant_table = dynamodb.Table('maria_history')
names_table = dynamodb.Table('slack_usernames')
channels_table = dynamodb.Table('channels_table')
meetings_table = dynamodb.Table('meetings_table')

#S3 Buckets 
image_bucket_name = 'mariaimagefolder-us'
docs_bucket_name = 'mariadocsfolder-us'

current_datetime = datetime.datetime.now()  
ai_temperature = 0.9

def lambda_handler(event, context):
    print(f"Event: {event}")   
    
    global stopwords
    global chat_id
    global user_id
    global thread_ts
    global audio_text
    global conversation

    stopwords = load_stopwords('english')
    has_image = ''
    audio_text = ''
    image_urls = []   
    has_image_urls = False

    rl = RouteLayer.from_json("routes_layer.json")
    
    if 'Records' in event and isinstance(event['Records'], list):
        payload = event['Records'][0]['body']
        body_str = json.loads(payload)['payload']['event']
        event_type = body_str['type']
        #print(f"Mail Body: {body_str}")
        system_text = prompts['system_text']
        email_instructions = prompts['email_instructions']

        # Get the current time as a Unix timestamp with fractional seconds    
        timestamp = time.time()
        
        # Format the timestamp as a string with 6 decimal places   
        thread_ts = f"{timestamp:.6f}"
        #print(f"Thread_ts: {thread_ts}")
        chat_id = "C06QL84KZGQ"
        user_id = "U02RPR2RMJS"
        
        user = get_users(user_id)
        user_name = user['real_name']
        display_name = user['display_name']
        display_name = display_name.replace(' ', '_').replace('.', '').strip()
        text = f"Email: {body_str} "
        #print(text)
        conversation = make_vision_conversation(system_text, email_instructions,  display_name, '', '', text)
        response_message = make_openai_vision_call(client, conversation)
        
    else:
        body_str = event.get('body', '{}')
        print(f"Body: {body_str}")
    
        # Parse the incoming event from Slack          
        slack_event = json.loads(event['body'])
        event_type = slack_event['event']['type']
        #print(event_type)
          
        #Initialize global variables        
        chat_id = slack_event['event']['channel']
        #print(f"Chat Id: {chat_id}")
        
        thread_ts = slack_event['event'].get('thread_ts', slack_event['event']['ts'])
    
        # Check if the 'user' key exists and has a value 
        if 'user' in slack_event['event'] and slack_event['event']['user']:
            user_id = slack_event['event']['user']
        else:
            # Stop execution if no user ID is found 
            print("No user ID found. Stopping execution.")
            return {
               'statusCode': 200,
               'body': json.dumps({'message': 'Ignored message.'})
            } # or use 'exit()' or 'sys.exit()' depending on the context  
    
        
        # Ignore messages from the bot itself  
        if 'event' in slack_event and 'bot_id' in slack_event['event']:
               return {
                   'statusCode': 200,
                   'body': json.dumps({'message': 'Ignored message from bot'})
               }
        
        # Retrieve the name of the user         
        user = get_users(user_id)
        user_name = user['real_name']
        display_name = user['display_name']
        display_name = display_name.replace(' ', '_').replace('.', '').strip()
        
        try:
            text = slack_event['event']['text']
        except KeyError:
            print(f"An error occured. No text found. slack_event: {json.dumps(slack_event)}") # Handle the error, e.g., log it and set a default value for text
            text = ""
        
        #Check for image in message 
        try:
            for file in slack_event["event"]["files"]:
                if file["mimetype"].startswith("image/"):
                    image_url = file["url_private"]
                    uploaded_url = upload_image_to_s3(image_url, image_bucket_name)
                    image_urls.append(uploaded_url)
                    
            #print(json.dumps(image_urls))
        except KeyError:
            image_urls = None    
    

        audio_urls = []
        speech_instruction = prompts['speech_instruction']

        # Check for audio in message     
        try:
            for file in slack_event["event"]["files"]:
                if file["mimetype"].startswith(("audio/", "video/")):
                    audio_url = file["url_private"]
                    audio_urls.append(audio_url)

            print(json.dumps(audio_urls))
            
            audio_text = transcribe_multiple_urls(audio_urls)
            # Check if audio_text is not empty  
            if audio_text:
                # Append speech_instruction to audio_text where audio_text is a list and speech_instruction is a string
                audio_text.append(speech_instruction)
                # Convert the array to a string with spaces in between each element
                combined_text = ' '.join(audio_text)
            print(f"Audio Msgs: {audio_text}")
            
        except KeyError:
            audio_urls = None
       
        size_limit_mb = 5  # Set the file size limit in MB
    
        # Check for documents in message
        try:
            application_files = []
            for file in slack_event["event"]["files"]:
                if file["mimetype"].startswith("application/") or file["mimetype"].startswith("text/"):
                    file_url = file["url_private"]
                    file_name = file.get("name", "Unnamed File")
                    #print(f"Mime Type: {file['mimetype']}")
                    file_size_mb = file.get("size", 1) / (1024 * 1024)
                    if file_size_mb > size_limit_mb:
                        application_files.append({"file_name": file_name, "content": f"File {file_name} is over the {size_limit_mb} MB limit. URL: {file_url}"})
                    else:
                        file_content = download_and_read_file(file_url, file["mimetype"])
                        application_files.append({"Message": "The user sent a file with this message. The contents of the file have been appended to this message.", "Filename": file_name, "content": file_content})
        
            print(json.dumps(application_files))   
        except KeyError:
            application_files = None
    
        # Check if there are attachments in the incoming message       
        if 'attachments' in slack_event['event']:
            attachments = slack_event['event']['attachments']
            
            # Extract content from attachments and append it to the text       
            for attachment in attachments:
                if 'text' in attachment:
                    forwarded_message_text = attachment['text']
                    # Append the forwarded message text along with a notice. 
                    text += f"\n\nForwarded Message:\n{forwarded_message_text}"
                    
        text += " " + " ".join(audio_text) if audio_text else ""
        text += " " + json.dumps(application_files) if application_files else ""

        route_name = ""
        if text:
            route_choice = rl(text)
            print(f'Route Choice: {route_choice}')
            route_name = route_choice.name

        # Retrieve the last N messages from the user and assistant     
        if route_name == 'chitchat':
            summary_len = 0
            full_text_len = 2
            system_text = prompts['instruct_basic']
            assistant_text = ""
        elif route_name == 'writing':
            system_text = prompts['system_text']
            assistant_text = prompts['assistant_text'] + " " + prompts['instruct_writing']
            summary_len = 10
            full_text_len = 10 
        elif route_name == 'odoo_erp':
            system_text = prompts['system_text']
            assistant_text = prompts['assistant_text'] + " " + prompts['odoo_search']
            summary_len = 2
            full_text_len = 5
            ai_temperature = 0.1   
        else:
            system_text = prompts['system_text']
            assistant_text = prompts['assistant_text'] + " " + prompts['odoo_search'] 
            summary_len = 10
            full_text_len = 10

        num_messages = summary_len + full_text_len  # or any other number you prefer

        # Extract and combine message history 
        user_msg_history, user_messages = (lambda x: (x[full_text_len:], x[:full_text_len]))(get_last_messages(user_table, chat_id, num_messages))
        #print(json.dumps(user_messages, default=decimal_default))
    
        asst_msg_history, assistant_messages = (lambda x: (x[full_text_len:], x[:full_text_len]))(get_last_messages(assistant_table, chat_id, num_messages))
        #print(json.dumps(asst_msg_history, default=decimal_default))   
    
        msg_history = user_msg_history + asst_msg_history
        msg_history.sort(key=lambda x: x['sort_key'])
        #print(json.dumps(msg_history, default=decimal_default))  
        
        msg_history_summary = [summarize_messages(msg_history)]
        #print(f"Message History Summary: {json.dumps(msg_history_summary, default=decimal_default)}")    
        
        # Combine and sort messages from user and assistant by sort_key
        all_messages = user_messages + assistant_messages
        all_messages.sort(key=lambda x: x['sort_key'])
        #print(json.dumps(all_messages, default=decimal_default))  
        
        try:
            maria_muted = manage_mute_status()[0]
            match_id = re.search(r"<@(\w+)>", slack_event['event']['text'])
            mentioned_user_id = match_id.group(1) if match_id else None
            print(f"Maria muted: {maria_muted}")
            # Check if Maria is muted and the mentioned user is not the specified ID
            if maria_muted and mentioned_user_id != "U05SSQR07RS":
                print("We are in the Maria is muted IF...")
                save_message(meetings_table, chat_id, text, "user", thread_ts, image_urls)
                return  # Exit the function after saving the message
        except Exception as e:
            print(f"An unexpected error occurred: {e}")            
        
        # Check if image_urls exists
        has_image_urls, all_image_urls = find_image_urls(all_messages)

        if image_urls or has_image_urls:
            # Prepare vision conversation and make API call
            conversation = make_vision_conversation(system_text, assistant_text, display_name, msg_history_summary, all_messages, text, image_urls)
            response_message = make_openai_vision_call(client, conversation)
        #elif audio_urls:
            # Prepare audio conversation and make API call
            #conversation = make_audio_conversation(system_text, assistant_text, display_name, msg_history_summary, all_messages, text, audio_urls)
            #response_message = make_openai_audio_call(client, conversation)
        else:
            # Prepare vision conversation without additional inputs and make API call
            conversation = make_vision_conversation(system_text, assistant_text, display_name, msg_history_summary, all_messages, text)
            response_message = make_openai_vision_call(client, conversation)


    #print(f"save_message called with user_table: {user_table}, chat_id: {chat_id}, text: '{text}', role: 'user', thread_ts: {thread_ts}, image_urls: {image_urls}")        
    # Save the user's message to DynamoDB 
    save_message(user_table, chat_id, text, "user", thread_ts, image_urls)

    # Define available functions    
    available_functions = {
        "browse_internet": browse_internet,
        "google_search": google_search,
        "get_coordinates": get_coordinates,
        "get_weather_data": get_weather_data,
        "get_message_by_sort_id": get_message_by_sort_id,
        "get_messages_in_range": get_messages_in_range,
        "send_slack_message": send_slack_message,
        "send_audio_to_slack": send_audio_to_slack,
        "send_file_to_slack": send_file_to_slack,        
        "get_users": get_users,
        "get_channels": get_channels,
        "send_as_pdf": send_as_pdf,
        "list_files": list_files,
        "ask_openai_o1": ask_openai_o1,
        "get_embedding": get_embedding,
        "manage_mute_status": manage_mute_status,
        "get_models": get_models,
        "get_fields": get_fields,
        "create_record": create_record,
        "fetch_records": fetch_records,
        "update_record": update_record,
        "delete_records": delete_records,
        "print_record": print_record
    }    

    while True:
        # Handle message content if present
        if isinstance(response_message, dict):
            has_content = response_message.get("content")
            has_audio = response_message.get("audio")
            has_tool_calls = response_message.get("tool_calls")
        else:
            has_content = getattr(response_message, "content", None)
            has_audio = getattr(response_message, "audio", None)
            has_tool_calls = getattr(response_message, "tool_calls", None)

        if has_content or has_audio:
            handle_message_content(response_message, event_type, thread_ts, chat_id)

        # Check and process tool calls
        if has_tool_calls:
            conversation_with_tool_responses = handle_tool_calls(
                response_message, available_functions, chat_id, conversation, thread_ts
            )

            # Determine the type of call to make based on the presence of image or audio URLs
            if image_urls or has_image_urls:
                # Vision API call
                response_message = make_openai_vision_call(client, conversation_with_tool_responses)
            #elif audio_urls:
                # Audio API call
                #response_message = make_openai_audio_call(client, conversation_with_tool_responses)
            else:
                # Default to Vision API if neither images nor audio are present
                response_message = make_openai_vision_call(client, conversation_with_tool_responses)
        else:
            break  # Exit the loop if there are no tool calls


def make_text_conversation(system_text, assistant_text, display_name, msg_history_summary, all_messages, text):
    # Initialize conversation array         
    conversation = [
        { "role": "system", "content": system_text },
        { "role": "assistant", "content": assistant_text}        
    ]

    conversation.append({ 
        "role": "assistant", 
        "content": (json.dumps(msg_history_summary, default=decimal_default))
    })
    
    # Add them to the conversation array   
    for message in all_messages:
        conversation.append({
            "role": message['role'],
            "content": message['message']
        }) 

    conversation.append({
        "role": "assistant",
        "content": "All the previous messages are a trail of the message history to aid your understanding of the conversation. The next message is the current request from the user."
    })    
    
    if display_name is not None and display_name != "":
        conversation.append({
            "role": "user",
            "content": text,
            "name": display_name
        })
    else:
        conversation.append({
            "role": "user",
            "content": text
        })

    print(json.dumps(conversation))         
    
    return conversation 


def make_vision_conversation(system_text, assistant_text, display_name, msg_history_summary, all_messages, text, image_urls=None): 
    # Initialize conversation array  
    conversation = [
        { "role": "system", "content": [{"type": "text", "text": system_text}] },
        { "role": "assistant", "content": [{"type": "text", "text": assistant_text}]}
    ]

    conversation.append({ 
        "role": "assistant", 
        "content": [{"type": "text", "text": json.dumps(msg_history_summary, default=decimal_default)}]
    })
    
    # Add them to the conversation array  
    for message in all_messages:
        # Start with the text content
        content = [{"type": "text", "text": message['message']}]
        
        # Check if there are image_urls in the message and add them to the content
        if "image_urls" in message:
            for url in message["image_urls"]:
                if url:  # Ensure the URL is not empty
                    content.append({"type": "image_url", "image_url": {"url": url}})
        
        # Add the constructed content to the conversation
        conversation.append({
            "role": message['role'],
            "content": content  
        })

    conversation.append({
        "role": "assistant",
        "content": "All the previous messages are a trail of the message history to aid your understanding of the conversation. The next message is the current request from the user."
    })    

    # Add final user message, optionally with image  
    user_content = [{"type": "text", "text": text}]

    # Assuming user_content is already defined and is a list  
    if image_urls is not None:
        for url in image_urls:
            if url:  # Check if the URL is not empty
                user_content.append({"type": "image_url", "image_url": {"url": url}})

    if display_name is not None and display_name != "":
        conversation.append({
            "role": "user",
            "content": user_content,
            "name": display_name
        })
    else:
        conversation.append({
            "role": "user",
            "content": user_content
        })

    print(json.dumps(conversation))    
    
    return conversation


def make_audio_conversation(system_text, assistant_text, display_name, msg_history_summary, all_messages, text, audio_urls=None):
    """
    Constructs the conversation payload for audio inputs by encoding audio files in base64.
    """
    # Initialize conversation array
    conversation = [
        {"role": "system", "content": [{"type": "text", "text": system_text}]},
        {"role": "assistant", "content": [{"type": "text", "text": assistant_text}]},
    ]

    conversation.append({
        "role": "assistant",
        "content": [{"type": "text", "text": json.dumps(msg_history_summary)}],
    })

    # Add historical messages
    for message in all_messages:
        content = [{"type": "text", "text": message["message"]}]
        conversation.append({
            "role": message["role"],
            "content": content,
        })

    # Add final system message
    conversation.append({
        "role": "assistant",
        "content": "All the previous messages are a trail of the message history to aid your understanding of the conversation. The next message is the current request from the user.",
    })

    # Add the user's current message
    user_content = [{"type": "text", "text": text}]
    if audio_urls is not None:
        for url in audio_urls:
            if url:
                # Download and encode the audio file
                response = requests.get(url)
                response.raise_for_status()
                m4a_data = response.content

                # Step 2: Convert M4A data to WAV in memory
                wav_data = convert_to_wav_in_memory(m4a_data)

                encoded_string = base64.b64encode(wav_data).decode("utf-8")
                #print(f"Encoded Audio: {encoded_string}")
                user_content.append({
                    "type": "input_audio",
                    "input_audio": {
                        "data": encoded_string,
                        "format": "wav"
                    }
                })

    if display_name:
        conversation.append({
            "role": "user",
            "content": user_content,
            "name": display_name,
        })
    else:
        conversation.append({
            "role": "user",
            "content": user_content,
        })
    
    print(f"Conversation: {json.dumps(conversation)}")

    return conversation


def handle_message_content(response_message, event_type, thread_ts, chat_id):
    """
    Handles both text and audio responses from OpenAI.
    Works with both dictionary and ChatCompletionMessage object formats.
    """
    global audio_text
    print(f"Response Message: {response_message}")

    # Check format type
    if isinstance(response_message, dict):
        # Dictionary format
        if "audio" in response_message:
            assistant_reply = response_message["audio"].get("transcript", "")
            audio_file_path = response_message["audio"].get("file_path", "")
            audio_text = assistant_reply
        else:
            assistant_reply = response_message["content"]
            #audio_text = None
    else:
        # ChatCompletionMessage format
        if getattr(response_message, "audio", None):
            assistant_reply = response_message.audio.get("transcript", "")
            audio_file_path = response_message.audio.get("file_path", "")
            audio_text = assistant_reply
        else:
            assistant_reply = response_message.content
            #audio_text = None

    # Rest of function remains the same
    save_message(assistant_table, chat_id, assistant_reply, "assistant", thread_ts)

    if event_type in ['app_mention', 'New Email']:
        if audio_text:
            send_audio_to_slack(assistant_reply, chat_id, thread_ts)
        else:
            send_slack_message(assistant_reply, chat_id, thread_ts)
    else:
        if audio_text:
            send_audio_to_slack(assistant_reply, chat_id, None)
        else:
            send_slack_message(assistant_reply, chat_id, None)


def handle_tool_calls(response_message, available_functions, chat_id, conversations, thread_ts):
    tool_calls = response_message.tool_calls

    response_message_dict = serialize_chat_completion_message(response_message)
    conversations.append(response_message_dict)
    
    def process_tool_call(tool_call):
        try:
            function_name = tool_call.function.name
            function_to_call = available_functions.get(function_name)

            if function_to_call:
                function_args = json.loads(tool_call.function.arguments)
                function_response = function_to_call(**function_args)

                if not isinstance(function_response, str):
                    function_response = json.dumps(function_response, default=decimal_default)
                    print(f"Function Response: {function_response}")
                    if function_name != "google_search":
                        save_message(assistant_table, chat_id, function_response, "assistant", thread_ts)
                    
                return {
                    "tool_call_id": tool_call.id,
                    "role": "tool",
                    "name": function_name,
                    "content": function_response,
                }
        except Exception as e:
            logging.error(f"Error processing tool call {tool_call.id}: {str(e)}")
            return {
                "tool_call_id": tool_call.id,
                "role": "tool",
                "name": function_name,
                "content": json.dumps({"error": str(e)}),
            }
        return None

    # Use ThreadPoolExecutor to process tool calls in parallel
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_tool_call = {executor.submit(process_tool_call, tool_call): tool_call for tool_call in tool_calls}
        
        for future in concurrent.futures.as_completed(future_to_tool_call):
            tool_call_result = future.result()
            if tool_call_result:
                conversations.append(tool_call_result)

    return conversations


def make_openai_api_call(client, conversations):
    try:
        # Prepare the API call 
        response = client.chat.completions.create(
            temperature=ai_temperature,
            model="gpt-4o-2024-11-20",
            messages=conversations,
            max_tokens=5500,
            tools=tools
        )
        print(response)
        return response.choices[0].message
    except Exception as e:
        print(f"An error occurred during the OpenAI API call: {e}")
        return None


def make_openai_vision_call(client, conversations):
    try:
        # Prepare the API call   
        response = client.chat.completions.create(
            model="gpt-4o-2024-11-20",
            messages=conversations,
            max_tokens=5500,
            tools=tools
        )
        print(response)
        return response.choices[0].message
    except Exception as e:
        print(f"An error occurred during the OpenAI Vision API call: {e}")
        return None


def make_openai_audio_call(client, conversations):
    """
    Sends the constructed conversation to OpenAI's chat completion API for audio processing.

    Args:
        client: OpenAI client instance.
        conversations (list): The list of messages for the conversation.

    Returns:
        dict: OpenAI message object with audio file path and transcript in `message.audio`.
    """
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {openai_api_key}",
    }
    payload = {
        "model": "gpt-4o-audio-preview-2024-12-17",
        "modalities": ["text", "audio"],
        "audio": {"voice": "shimmer", "format": "mp3"},
        "messages": conversations,
        "max_tokens": 5500,
    }

    try:
        response = requests.post(url, headers=headers, data=json.dumps(payload))
        response.raise_for_status()

        result = response.json()
        print(f"Openai Response: {result}")
        
        message = result["choices"][0]["message"]

        # Check if audio is present
        if "audio" in message:
            # Decode audio and write to a temporary file
            audio_data = base64.b64decode(message["audio"]["data"])
            with tempfile.NamedTemporaryFile(delete=False, suffix=".mp3") as tmp_audio_file:
                tmp_audio_file.write(audio_data)
                audio_file_path = tmp_audio_file.name

            # Update the audio object with the file path
            message["audio"]["file_path"] = audio_file_path
            del message["audio"]["data"]

        print(f"Openai Modified Response: {message}")

        return message

    except requests.exceptions.RequestException as e:
        # Handle HTTP errors
        if hasattr(e, 'response'):
            print(f"Status code: {e.response.status_code}")
            print(f"Error response: {e.response.json()}")  # OpenAI sends error details in JSON
            print(f"Headers: {e.response.headers}")
        else:
            print(f"Error with no response: {str(e)}")
        return None
    except KeyError as e:
        # Handle missing keys in the JSON response
        print(f"KeyError: Missing expected key in response: {e}")
        return None
    except Exception as e:
        # Handle unexpected errors
        print(f"Unexpected error: {str(e)}")
        return None


def ask_openai_o1(prompt):
    print(f'OpenAI o1 Prompt: {prompt}')
    message = [
        {
            "role": "user",
            "content": prompt
        }       
    ]

    try:
        # Prepare the API call   
        response = client.chat.completions.create(
            model="o1",
            messages=message
        )
        print(f'OpenAI o1: {response}')
        
        # Extract the actual message content using dot notation
        response_message_content = response.choices[0].message.content

        # Return the serialized content
        return json.dumps(response_message_content, default=decimal_default)
    except Exception as e:
        print(f"An error occurred during the OpenAI o1 API call: {e}")
        return None


def save_message(table, chat_id, text, role, thread=None, image_urls=None):
    timestamp = int(time.time())
    sort_key = timestamp  # use timestamp as sort key
    ttl = timestamp + 20 * 24 * 60 * 60  # 20 days

    item = {
        'chat_id': chat_id,
        'sort_key': sort_key,
        'message': text,
        'role': role,  # Existing line to save the role of the message  
        'ttl': ttl
    }

    # Add thread to the item if it's provided
    if thread is not None:
        item['thread'] = thread
    
    # Add image urls to the item if they are provided
    if image_urls is not None:
        item['image_urls'] = image_urls
    
    #print(f"Item : {json.dumps(item)}")    

    table.put_item(Item=item)


def get_last_messages(table, chat_id, num_messages):
    response = table.query(
        KeyConditionExpression=Key('chat_id').eq(chat_id),
        Limit=num_messages,
        ScanIndexForward=False  # get the latest messages first
    )
    messages = response['Items'] if 'Items' in response else []
    return messages  # return the whole item, not just the message
    
    
def get_message_by_sort_id(role, chat_id, sort_id):
    # Determine the appropriate table based on the role 
    if role == 'user':
        table = user_table  # Assuming user_table is already defined
    elif role == 'assistant':
        table = assistant_table  # Assuming assistant_table is already defined 
    else:
        return None  # or handle the error appropriately   

    response = table.query(
        KeyConditionExpression=Key('chat_id').eq(chat_id) & Key('sort_key').eq(sort_id),
        Limit=1  # We expect only one item
    )
    message = response['Items'][0] if 'Items' in response and response['Items'] else None
    return message  # return the whole item, not just the message 


def get_messages_in_range(chat_id, start_sort_id, end_sort_id):
    
    # Function to query a table
    def query_table(table):
        return table.query(
            KeyConditionExpression=Key('chat_id').eq(chat_id) & Key('sort_key').between(start_sort_id, end_sort_id)
        )['Items']

    # Retrieve messages from both tables  
    user_messages = query_table(user_table)
    assistant_messages = query_table(assistant_table)

    # Combine and sort messages from user and assistant by sort_key
    all_messages = user_messages + assistant_messages
    all_messages.sort(key=lambda x: x['sort_key'])
    all_messages_summary = [summarize_messages(all_messages)]
    print(f"All Messages Summary: {json.dumps(all_messages_summary, default=decimal_default)}")
    return all_messages_summary


def send_slack_message(message, channel, ts=None):
    # Slack Bot URL
    url = "https://slack.com/api/chat.postMessage"

    #text_message = convert_markdown_to_slack(message)
    text_message = BeautifulSoup(markdown2.markdown(message), 'html.parser').get_text()
    slack_blocks = convert_to_slack_blocks(message)
    #print(f"Text Message: {text_message}") 
    print(f"Slack Blocks: {slack_blocks}") 
    
    # Message data
    data = {
        'channel': channel,
        'text': message,
        'thread_ts' : ts,
        'blocks' : slack_blocks
    }
    
    # HTTP headers   
    headers = {
        'Content-Type': 'application/json; charset=utf-8',
        'Authorization': 'Bearer ' +  slack_bot_token
    }

    response = requests.post(url, headers=headers, json=data)
    #print(response)

    return response.json()


def send_audio_to_slack(text, chat_id=None, ts=None):
    """
    Converts text to speech and uploads the audio file to Slack using the external upload API.
    
    Parameters:
    - text: The text to be converted to speech
    - chat_id: The ID of the Slack channel where the file will be uploaded (optional)
    - ts: Thread timestamp to attach the file to (optional)
    
    Returns:
    - dict: The JSON response from the files.completeUploadExternal API
    """
    # Use the global chat_id if it's not passed as an argument
    if chat_id is None:
        chat_id = globals()['chat_id']
    
    # Convert text to speech and get file details
    file_path = text_to_speech(text, file_suffix=".mp3")
    file_size = os.path.getsize(file_path)
    file_name = os.path.basename(file_path)
    
    # Step 1: Get upload URL
    headers = {
        "Authorization": f"Bearer {slack_bot_token}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    url_params = {
        "filename": file_name,
        "length": file_size
    }
    
    response = requests.get(
        "https://slack.com/api/files.getUploadURLExternal",
        headers=headers,
        params=url_params
    )
    
    if not response.ok or not response.json().get("ok"):
        error = response.json().get("error", "Unknown error")
        print(f"Error getting upload URL: {error}")
        return response.json()
    
    upload_url = response.json()["upload_url"]
    file_id = response.json()["file_id"]
    
    # Step 2: Upload file to the provided URL
    with open(file_path, "rb") as file:
        files = {
            "file": (file_name, file, "audio/mpeg")
        }
        upload_response = requests.post(upload_url, files=files)
    
    if not upload_response.ok:
        print(f"Error uploading file: {upload_response.status_code}")
        return {"ok": False, "error": "upload_failed"}
    
    # Step 3: Complete the upload
    complete_data = {
        "files": [{
            "id": file_id,
            "title": "Audio Response"
        }]
    }
    
    # Add optional parameters if provided
    if chat_id:
        complete_data["channel_id"] = chat_id
    if ts:
        complete_data["thread_ts"] = ts
    
    complete_response = requests.post(
        "https://slack.com/api/files.completeUploadExternal",
        headers={
            "Authorization": f"Bearer {slack_bot_token}",
            "Content-Type": "application/json"
        },
        json=complete_data
    )
    
    if complete_response.ok and complete_response.json().get("ok"):
        print(f"File uploaded successfully: {file_id}")
    else:
        error = complete_response.json().get("error", "Unknown error")
        print(f"Error completing upload: {error}")
    
    return complete_response.json()


def send_file_to_slack(file_path, chat_id, title, ts=None):
    """
    Uploads a file to a specified Slack channel using the external upload API.
    If the file_path is a URL, downloads the file first.
    
    Parameters:
    - file_path: The path to the file or a URL
    - chat_id: The ID of the Slack channel where the file will be uploaded
    - title: The title of the file
    - ts: The thread timestamp (optional)
    
    Returns:
    - dict: The JSON response from the files.completeUploadExternal API
    """
    # Check if file_path is a URL
    parsed_url = urlparse(file_path)
    is_url = parsed_url.scheme in ('http', 'https')
    
    if is_url:
        # Download the file into a temporary file
        response = requests.get(file_path)
        response.raise_for_status()  # Ensure the download succeeded
        suffix = os.path.splitext(parsed_url.path)[1]  # Extract the file extension
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            tmp_file.write(response.content)
            file_to_upload = tmp_file.name
    else:
        file_to_upload = file_path

    try:
        # Get file details
        file_size = os.path.getsize(file_to_upload)
        file_name = os.path.basename(file_to_upload)
        
        # Step 1: Get upload URL
        headers = {
            "Authorization": f"Bearer {slack_bot_token}",
            "Content-Type": "application/x-www-form-urlencoded"
        }
        
        url_params = {
            "filename": file_name,
            "length": file_size
        }
        
        response = requests.get(
            "https://slack.com/api/files.getUploadURLExternal",
            headers=headers,
            params=url_params
        )
        
        if not response.ok or not response.json().get("ok"):
            error = response.json().get("error", "Unknown error")
            print(f"Error getting upload URL: {error}")
            return response.json()
        
        upload_url = response.json()["upload_url"]
        file_id = response.json()["file_id"]
        
        # Step 2: Upload file to the provided URL
        with open(file_to_upload, "rb") as file:
            # Determine content type based on file extension
            file_extension = os.path.splitext(file_to_upload)[1].lower()
            content_type = mimetypes.guess_type(file_to_upload)[0] or "application/octet-stream"
            
            files = {
                "file": (file_name, file, content_type)
            }
            upload_response = requests.post(upload_url, files=files)
        
        if not upload_response.ok:
            print(f"Error uploading file: {upload_response.status_code}")
            return {"ok": False, "error": "upload_failed"}
        
        # Step 3: Complete the upload
        complete_data = {
            "files": [{
                "id": file_id,
                "title": title
            }],
            "channel_id": chat_id
        }
        
        if ts:
            complete_data["thread_ts"] = ts
        
        complete_response = requests.post(
            "https://slack.com/api/files.completeUploadExternal",
            headers={
                "Authorization": f"Bearer {slack_bot_token}",
                "Content-Type": "application/json"
            },
            json=complete_data
        )
        
        if complete_response.ok and complete_response.json().get("ok"):
            print(f"File uploaded successfully: {file_id}")
        else:
            error = complete_response.json().get("error", "Unknown error")
            print(f"Error completing upload: {error}")
        
        return complete_response.json()
        
    finally:
        if is_url:
            # Remove the temporary file
            os.unlink(file_to_upload)


"""
def send_file_to_slack(file_path, chat_id, title, ts=None):
    
    #Uploads a file to a specified Slack channel. If the file_path is a URL, downloads the file first. 

    #Parameters:
    #- file_path: The path to the file or a URL.
    #- chat_id: The ID of the Slack channel where the file will be uploaded.
    #- title: The title of the file.
    #- ts: The thread timestamp (optional).
    
    # Check if file_path is a URL
    parsed_url = urlparse(file_path)
    is_url = parsed_url.scheme in ('http', 'https')

    if is_url:
        # Download the file into a temporary file
        response = requests.get(file_path)
        response.raise_for_status()  # Ensure the download succeeded
        suffix = os.path.splitext(parsed_url.path)[1]  # Extract the file extension
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            tmp_file.write(response.content)
            file_to_upload = tmp_file.name
    else:
        file_to_upload = file_path

    try:
        file_extension = os.path.splitext(file_to_upload)[1].lower()
        file_type = file_extension[1:]  # Remove the leading dot

        headers = {
            "Authorization": f"Bearer {slack_bot_token}"
        }

        data = {
            "channels": chat_id,
            "title": title,
            "filetype": file_type,
            "thread_ts": ts
        }

        files = {
            "file": open(file_to_upload, "rb")
        }

        response = requests.post("https://slack.com/api/files.upload", headers=headers, data=data, files=files)

        if response.status_code == 200 and response.json().get("ok"):
            print(f"File uploaded successfully: {response.json()['file']['name']}")
        else:
            print(f"Error uploading file: {response.json().get('error', 'Unknown error')}")

        return response.json()
    finally:
        if is_url:
            # Remove the temporary file
            os.unlink(file_to_upload)
"""
   
def message_to_json(message_string):
    message_parts = re.split(r'```json|```', message_string)
    
    message = {
        "assistant": f"{message_parts[0].strip()} {message_parts[2].strip()}",
        "blocks": json.loads(message_parts[1])['blocks']
    }
    
    return json.dumps(message)
    

def get_slack_user_name(user_id):
    """
    Function to get the name of the user who sent a message in Slack.
    
    Parameters: 
    token (str): The Slack bot token.
    user_id (str): The user ID of the message sender.
    
    Returns:
    str: The real name of the user if found, otherwise an error message.
    """
    url = 'https://slack.com/api/users.info'
    headers = {
        'Authorization': f'Bearer {slack_bot_token}',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    params = {
        'user': user_id
    }
    
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        user_info = response.json()
        if user_info.get('ok'):
            # Extracting the user's real name from the response
            user_name = user_info['user']['profile']['real_name']
            display_name = user_info['user']['profile']['display_name']
            email = user_info['user']['profile']['email']
            return {"User Name" : user_name, "Display Name" : display_name, "Email" : email}
        else:
            # Return the error message if the 'ok' field is False
            return user_info.get('error', 'Unknown error occurred')
    else:
        return f'HTTP Error: {response.status_code}'


def get_users(user_id=None):
    try:
        if user_id:
            # Retrieve a single user
            response = names_table.get_item(Key={'user_id': user_id})
            item = response.get('Item', None)
            if item:
                #print("Item found:", item)
                return item
            else:
                #print(user_id, " not found. Retrieving name from slack.")
                update_slack_users()
                response = names_table.get_item(Key={'user_id': user_id})
                item = response.get('Item', None)
                if item:
                    #print("Item found:", item)
                    return item
                else:
                    print(user_id, " still not found after update.")
                    return None
        else:
            # Retrieve all users
            response = names_table.scan()
            items = response.get('Items', [])
            #print(items)
            return items
    except Exception as e:
        print(f"Error: {e}")
        return None

def update_slack_users():
    url = 'https://slack.com/api/users.list'
    headers = {
        'Authorization': f'Bearer {slack_bot_token}',
        'Content-Type': 'application/x-www-form-urlencoded'
    }

    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        users_list = response.json().get('members', [])
        #print(users_list)
        for user in users_list:
            if not user.get('deleted', True) and not user.get('is_bot', True) and not user.get('is_app_user', True):
                user_id = user.get('id')
                try:
                    response = names_table.get_item(Key={'user_id': user_id})
                    item = response.get('Item', None)
                    if item:
                        print(f"User {user_id} already exists in the database.")
                        # Check if all keys are available
                        missing_keys = [key for key in ['real_name', 'display_name', 'email'] if key not in item]
                        if missing_keys:
                            print(f"Updating user {user_id} with missing keys: {missing_keys}")
                            for key in missing_keys:
                                item[key] = user.get('profile', {}).get(key, '')
                            names_table.put_item(Item=item)
                            print(f"User {user_id} updated successfully.")
                    else:
                        print(f"Adding user {user_id} to the database.")
                        user_data = {
                            'user_id': user_id,
                            'real_name': user.get('profile', {}).get('real_name', ''),
                            'display_name': user.get('profile', {}).get('display_name', ''),
                            'email': user.get('profile', {}).get('email', '')
                        }
                        names_table.put_item(Item=user_data)
                        print(f"User {user_id} added successfully.")
                except Exception as e:
                    print(f"Error processing user {user_id}: {e}")
    else:
        print(f"Failed to retrieve users list: HTTP {response.status_code}")


def serialize_chat_completion_message(message):
    """
    Converts a ChatCompletionMessage object into a serializable dictionary.
    """
    message_dict = {
        "content": message.content,
        "role": message.role
    }

    # Serialize the function_call if it exists
    if message.function_call:
        message_dict["function_call"] = {
            "name": message.function_call.name,
            "arguments": message.function_call.arguments  # Assuming arguments are in JSON format
        }

    # Serialize tool_calls if they exist
    if message.tool_calls:
        message_dict["tool_calls"] = []
        for tool_call in message.tool_calls:
            message_dict["tool_calls"].append({
                "id": tool_call.id,
                "function": {
                    "name": tool_call.function.name,
                    "arguments": tool_call.function.arguments  # Assuming arguments are in JSON format
                },
                "type": tool_call.type
            })

    return message_dict


def load_stopwords(file_path):
    """
    Load stopwords from a given file.

    :param file_path: Path to the stopwords file.
    :return: Set of stopwords.
    """
    with open(file_path, 'r') as file:
        stopwords = set(file.read().splitlines())
    return stopwords

def rank_sentences(text, stopwords, max_sentences=10):
    """
    Rank sentences in the text based on word frequency, returning top 'max_sentences' sentences.   
    """
    word_frequencies = {}
    for word in word_tokenize(text.lower()):
        if word.isalpha() and word not in stopwords:  # Consider only alphabetic words
            word_frequencies[word] = word_frequencies.get(word, 0) + 1

    sentence_scores = {}
    sentences = sent_tokenize(text)
    for sent in sentences:
        for word in word_tokenize(sent.lower()):
            if word in word_frequencies and len(sent.split(' ')) < 30:
                sentence_scores[sent] = sentence_scores.get(sent, 0) + word_frequencies[word]

    sorted_sentences = sorted(sentence_scores, key=sentence_scores.get, reverse=True)
    summary_sentences = sorted_sentences[:max_sentences]
    
    # Add a full stop at the end of each sentence if it doesn't already end with one 
    summary = ' '.join([s if s.endswith('.') else f'{s}.' for s in summary_sentences])

    return summary


def summarize_record(record, stopwords):
    """
    Summarize a single message record while maintaining key points and brevity.
    """
    # Record format: 'sort_key: chat_id: role: message'
    parts = record.split(':', 3)
    sort_key, chat_id, role, message = parts[0], parts[1], parts[2], parts[3].strip()
    summarized_message = rank_sentences(message, stopwords, max_sentences=5)
    return {
        'sort_key': sort_key,
        'chat_id': chat_id,
        'role': role,
        'message_summary': summarized_message
    }


def summarize_messages(data):
    """
    Summarize messages from a dictionary and return a dictionary with the summarized conversation. 
    """
    #stopwords_file = 'english'
    #stopwords = load_stopwords(stopwords_file)
    
    # Create play-like records
    records = [f"{item['sort_key']}: {item['chat_id']}: {item['role']}: {item['message']}" for item in data]
    
    # Summarize each record
    summarized_records = [summarize_record(record, stopwords) for record in records]
    #print(json.dumps(summarized_records, default=decimal_default)) 
    
    return summarized_records


# Function to convert non-serializable types for JSON serialization  
def decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, ObjectId):
        return str(obj)  # Convert ObjectId to string
    raise TypeError("Unserializable object {} of type {}".format(obj, type(obj)))
    

def convert_floats_to_decimals(obj):
    if isinstance(obj, float):
        return decimal.Decimal(str(obj))
    elif isinstance(obj, dict):
        return {k: convert_floats_to_decimals(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_floats_to_decimals(v) for v in obj]
    return obj


def convert_markdown_to_slack(content):
    # Convert bold text
    content = re.sub(r"\*\*(.*?)\*\*", r"*\1*", content)

    # Convert italic text
    content = re.sub(r"\_(.*?)\_", r"_\1_", content)

    # Convert LaTeX (or similar) to Slack code format
    content = re.sub(r"\\\( (.*?) \\\)", r"`\1`", content)

    return content
    

def convert_to_slack_blocks(markdown_text):
    lines = markdown_text.split('\n')
    blocks = []
    current_section = []
    in_code_block = False
    in_blockquote = False

    def clean_heading_text(text):
        """Helper function to clean heading text by removing leading/trailing whitespace and asterisks"""
        text = text.strip()  # Remove whitespace
        # Remove leading/trailing asterisks
        while text.startswith('*') and text.endswith('*'):
            text = text[1:-1].strip()
        return text

    for line in lines:
        if in_code_block:
            # Check for the end of the code block
            if line.startswith("```"):
                # Add the current code block as a section
                blocks.append({
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": "```\n" + "\n".join(current_section) + "\n```",
                    }
                })
                current_section = []
                in_code_block = False
            else:
                # If we're inside a code block, append the line without modifying it
                current_section.append(line)
            continue
        
        # Handle horizontal rules
        if line.strip() in ('***', '---', '___'):
            if current_section:
                blocks.append({
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": "\n".join(current_section)
                    }
                })
                current_section = []
            blocks.append({"type": "divider"})
            continue
        
        # Handle the start of a code block
        if line.startswith("```"):
            # Flush any current text as a section
            if current_section:
                blocks.append({
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": "\n".join(current_section)
                    }
                })
                current_section = []
            in_code_block = True
            current_section.append(line)  # Include the starting backticks
            continue
        
        # Replace Markdown bold with Slack's bold syntax outside code blocks
        line = line.replace('**', '*')

        # Handle blockquotes
        if line.startswith('>'):
            if not in_blockquote:
                in_blockquote = True
                if current_section:
                    blocks.append({
                        "type": "section",
                        "text": {
                            "type": "mrkdwn",
                            "text": "\n".join(current_section)
                        }
                    })
                    current_section = []
            current_section.append(line[1:].strip())  # Remove '>' and add line
        elif in_blockquote and not line.startswith('>'):
            # End of blockquote
            in_blockquote = False
            blocks.append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "\n".join(current_section)
                }
            })
            current_section = []

        # Headers and regular text
        elif line.lstrip().startswith('#'):
            if current_section:
                blocks.append({
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": "\n".join(current_section)
                    }
                })
                current_section = []
            
            # Count the number of # symbols and remove them
            heading_level = 0
            line_stripped = line.lstrip()
            for char in line_stripped:
                if char == '#':
                    heading_level += 1
                else:
                    break
            
            # Extract the heading text, remove leading/trailing whitespace and asterisks
            heading_text = clean_heading_text(line_stripped[heading_level:])
            
            # Use progressively smaller markers as the heading level increases
            size_markers = ['◆', '🔷', '🔹', '▪️', '▫️', '·']  # For H1 through H6
            marker = size_markers[min(heading_level - 1, len(size_markers) - 1)]
            
            # All headings are now section blocks with bold text
            blocks.append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"{marker} *{heading_text}*"
                }
            })
        elif line.strip():
            current_section.append(line)

        # Empty lines and end of blockquotes
        elif current_section:
            blocks.append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "\n".join(current_section)
                }
            })
            current_section = []
            in_blockquote = False

    # Add the last section if there is any text left
    if current_section:
        blocks.append({
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": "\n".join(current_section)
            }
        })

    return json.dumps(blocks)


def latex_to_slack(latex_str):
    """
    Convert a LaTeX string to a Slack-friendly format.      
    
    Args:
    latex_str (str): A string containing LaTeX commands. 

    Returns:
    str: A string formatted for Slack.
    """
    # Replace common LaTeX commands with Slack-friendly equivalents     
    replacements = {
        '\\times': '×',  # Multiplication
        '\\frac': '/',   # Fractions
        '\\sqrt': '√',   # Square root
        '^': '**',       # Exponentiation
        '_': '~'         # Subscript
    }

    slack_str = latex_str
    for latex, slack in replacements.items():
        slack_str = slack_str.replace(latex, slack)

    return slack_str
    
    
def get_embedding(text, model="text-embedding-ada-002"):
    text_cleaned = text.replace("\n", " ")
    embedding = client.embeddings.create(input=[text_cleaned], model=model).data[0].embedding
    return {"text": text_cleaned, "embedding": embedding}
    
    

def send_to_sqs(data, queue_url):
    sqs = boto3.client('sqs')
    payload = {
        'payload': data
    }
    
    sqs.send_message(
        QueueUrl=queue_url,
        MessageBody=json.dumps(payload)
    )

def invoke_lambda(data):
    lambda_client = boto3.client('lambda')
    payload = {
        'payload': data
    }
    response = lambda_client.invoke(
        FunctionName='cerenyiWebsearch',
        InvocationType='RequestResponse',  # Changed to synchronous invocation
        Payload=json.dumps(payload),
    )
    
    # To get the response payload   
    response_payload = json.loads(response['Payload'].read())
    return response_payload


def google_search(search_term, before=None, after=None, intext=None, allintext=None, and_condition=None, must_have=None):
    # Initialize API keys from environment variables
    custom_search_api_key = os.getenv('CUSTOM_SEARCH_API_KEY')
    custom_search_id = os.getenv('CUSTOM_SEARCH_ID')

    # Constructing the search term with advanced operators
    search_components = [search_term]

    # Implementing 'and' search operator
    if and_condition:
        search_components.append(search_term + " AND " + and_condition)

    # Implementing 'before' search operator (YYYY-MM-DD format)
    if before:
        search_components.append(f"before:{before}")

    # Implementing 'after' search operator (YYYY-MM-DD format)
    if after:
        search_components.append(f"after:{after}")

    # Implementing 'intext' search operator
    if intext:
        search_components.append(f"intext:{intext}")
    
    # Implementing 'allintext' search operator
    if allintext:
        search_components.append(f"allintext:{allintext}")
    
    # Implementing 'must_have' operator to require exact phrase match
    if must_have:
        search_components.append(f"\"{must_have}\"")  # Use of escaped quotes to handle Python string encapsulation

    # Join all components to form the final search query
    combined_search_term = ' '.join(search_components)
    url_encoded_search_term = quote_plus(combined_search_term)
    print(f'Search Term: {url_encoded_search_term}')

    # Build the search URL   
    search_url = f"https://www.googleapis.com/customsearch/v1?q={url_encoded_search_term}&cx={custom_search_id}&key={custom_search_api_key}"
    response = requests.get(search_url)
    results = response.json().get('items', [])
    print(json.dumps(results, default=decimal_default))

    web_links = []
    for result in results:
        web_links.append(result['link'])
    
    # Assuming get_web_pages is a coroutine to fetch web pages 
    web_content = asyncio.run(get_web_pages(web_links[:5]))
    print(json.dumps(web_content, default=decimal_default))  # Assuming decimal_default was a function for JSON serializing decimals   
    
    return web_content


def browse_internet(urls, full_text=False):
    web_pages = asyncio.run(get_web_pages(urls, full_text))
    print(web_pages)
    return web_pages


async def get_web_pages(urls, full_text=False, max_concurrent_requests=5):
    async with aiohttp.ClientSession() as session:
        semaphore = asyncio.Semaphore(max_concurrent_requests)
        tasks = [process_page(session, url, semaphore, full_text) for url in urls]
        results = await asyncio.gather(*tasks)
        
        # Flatten the list of results
        flattened_results = [item for sublist in results for item in sublist]
        print(json.dumps(flattened_results))
        
        return flattened_results


async def fetch_page(session, url, timeout=30):
    headers = {
        'User-Agent': random.choice(USER_AGENTS)
    }    
    try:
        async with session.get(url, headers=headers, proxy=proxy_url, timeout=timeout) as response:
            #print(f"Search Result: {response}")
            content_type = response.headers.get('Content-Type', '')
            if 'text' in content_type:
                encoding = response.charset or 'utf-8'
                return await response.text(encoding=encoding)
            elif 'application/pdf' in content_type or 'application/msword' in content_type or 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type:
                return await response.read(), content_type
            else:
                return None, content_type
    except asyncio.TimeoutError:
        print(f"Timeout error: {url} took too long to respond.")
        return f"Timeout error: {url} took too long to respond.", None
    except ClientConnectorSSLError:
        print(f"SSL handshake error: Failed to connect to {url}")
        return f"SSL handshake error: Failed to connect to {url}", None 

async def process_page(session, url, semaphore, full_text=False):
    async with semaphore:
        try:
            result = await fetch_page(session, url)
        except ClientError as e:
            logging.error(f"Client error occurred while fetching the page: {e}")
            return [{
                "type": "text",
                "text": {
                    'url': url,
                    'error': 'Failed to fetch page due to client error'
                }
            }]
        except Exception as e:
            logging.error(f"Unexpected error occurred: {e}")
            return [{
                "type": "text",
                "text": {
                    'url': url,
                    'error': 'An unexpected error occurred while fetching the page'
                }
            }]
        
        response_list = []

        #print(f'Raw Result: {result}')
        try:
            if isinstance(result, tuple):
                document_content, content_type = result
                if document_content is not None and content_type is not None:
                    try:
                        s3_url = upload_document_to_s3(document_content, content_type, url)
                        response_list.append({
                            "type": "text",
                            "text": {
                                'url': url,
                                'summary': 'This file contains additional information for your search. Send it to the user.',
                                's3_url': s3_url
                            }
                        })
                    except Exception as e:
                        logging.error(f"Failed to upload document to S3: {e}")
                        response_list.append({
                            "type": "text",
                            "text": {
                                'url': url,
                                'error': 'Failed to upload document to S3'
                            }
                        })
                else:
                    response_list.append({
                        "type": "text",
                        "text": {
                            'url': url,
                            'error': 'Unsupported content type'
                        }
                    })
            elif isinstance(result, str) and 'Timeout error' not in result:
                soup = BeautifulSoup(result, 'lxml')

                elements_to_extract = ['p', 'li', 'summary', 'div', 'span', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'blockquote', 'pre', 'td', 'th', 'a']

                text = ' '.join(element.get_text().strip() for element in soup.find_all(elements_to_extract))
                cleaned_text = clean_website_data(text)

                if full_text:
                    summary_or_full_text = rank_sentences(cleaned_text, stopwords, max_sentences=150)  # Placeholder for the rank_sentences function
                else:
                    try:
                        summary_or_full_text = rank_sentences(cleaned_text, stopwords, max_sentences=50)  # Placeholder for the rank_sentences function
                    except Exception as e:
                        logging.error(f"Failed to rank sentences: {e}")
                        summary_or_full_text = cleaned_text  # Fallback to full text if ranking fails

                author = soup.find('meta', {'name': 'author'})['content'] if soup.find('meta', {'name': 'author'}) else 'Unknown'
                date_published = soup.find('meta', {'property': 'article:published_time'})['content'] if soup.find('meta', {'property': 'article:published_time'}) else 'Unknown'

                links = []

                response_list.append({
                    "type": "text",
                    "text": {
                        'summary_or_full_text': summary_or_full_text,
                        'author': author,
                        'date_published': date_published,
                        'internal_links': links
                    }
                })

                images = soup.select('article img') + soup.select('figure img') + soup.select('section img')
                
                for img in images:
                    img_url = img.get('src')
                    if img_url:
                        # Skip data URIs and other non-HTTP/HTTPS sources
                        if img_url.startswith('data:'):
                            #logging.warning(f"Skipping data URI image: {img_url[:30]}...")  # Log a warning and skip data URIs
                            continue
                        if not img_url.startswith(('http://', 'https://')):
                            img_url = urljoin(url, img_url)
                        
                        try:
                            async with session.head(img_url) as img_response:
                                if img_response.status == 200 and int(img_response.headers.get('Content-Length', 0)) > 10240:
                                    response_list.append({
                                        "type": "image_url",
                                        "image_url": {
                                            'url': img_url
                                        }
                                    })
                        except ClientError as e:
                            logging.error(f"Failed to fetch image: {img_url} - ClientError: {e}")
                        except Exception as e:
                            logging.error(f"Failed to fetch image: {img_url} - Unexpected error: {e}")
            else:
                response_list.append({
                    "type": "text",
                    "text": {
                        'url': url,
                        'error': result
                    }
                })
        except Exception as e:
            logging.error(f"Error processing page: {e}")
            response_list.append({
                "type": "text",
                "text": {
                    'url': url,
                    'error': 'An error occurred while processing the page'
                }
            })

        return response_list


def upload_document_to_s3(document_content, content_type, document_url):
    document_extension = mimetypes.guess_extension(content_type) or '.bin'
    s3_object_name = f"Document_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}{document_extension}"

    # Upload to S3   
    s3_client = boto3.client('s3')
    s3_client.put_object(Body=document_content, Bucket=docs_bucket_name, Key=s3_object_name)

    # Construct the S3 URL  
    s3_url = f"https://{docs_bucket_name}.s3.amazonaws.com/{s3_object_name}"
    return s3_url


def upload_image_to_s3(image_url, bucket_name):
    # Download the image   
    headers = {
        'Authorization': f'Bearer {slack_bot_token}',
        'Content-Type': 'image/x-www-form-urlencoded'
    }
    response = requests.get(image_url, headers=headers)   

    if response.status_code != 200:
        raise Exception('Failed to download image')
        
    # Get the image content  
    image_content = BytesIO(response.content)

    file_extension = os.path.splitext(image_url)[1]
    s3_object_name = f"Maria_{chat_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}{file_extension}"

    # Upload to S3
    s3_client = boto3.client('s3')
    s3_client.upload_fileobj(image_content, bucket_name, s3_object_name)

    # Construct the S3 URL    
    s3_url = f"https://{bucket_name}.s3.amazonaws.com/{s3_object_name}"
    return s3_url


def transcribe_speech_from_memory(audio_stream):
    # Reset the stream's position to the beginning
    audio_stream.seek(0)

    # Create a temporary file to store the audio content
    with tempfile.NamedTemporaryFile(suffix='.m4a', delete=False) as tmp_file:
        tmp_file.write(audio_stream.read())
        tmp_file_path = tmp_file.name

    # Open the temporary file for reading
    with open(tmp_file_path, 'rb') as audio_file:
        # Create a transcription using OpenAI's Whisper model
        response = client.audio.transcriptions.create(
            model="whisper-1",
            file=audio_file
        )

    #print(f"OpenAI : {response.text}")

    # Remove the temporary file  
    os.remove(tmp_file_path)
    return response.text


def download_audio_to_memory(url):
    headers = {
        'Authorization': f'Bearer {slack_bot_token}',
        'Content-Type': 'audio/x-www-form-urlencoded'
    }    
    response = requests.get(url, headers = headers)
    if response.status_code != 200:
        raise Exception('Failed to download audio')
        
    # Check response headers
    #print(f"Response Content-Type: {response.headers.get('Content-Type')}")
    
    audio_content = BytesIO(response.content)
    
    # Print the size of the downloaded audio (in bytes)
    #print(f"Size of downloaded audio: {audio_content.getbuffer().nbytes} bytes") 
    
    response.raise_for_status()
    return audio_content


def process_url(url):
    try:
        audio_stream = download_audio_to_memory(url)
        transcription = transcribe_speech_from_memory(audio_stream)
        print(f"Success: Transcription completed for URL: {url}")
        return transcription
    except Exception as e:
        print(f"Failure: Could not process URL: {url}. Error: {str(e)}")
        return None


def transcribe_multiple_urls(urls):
    results = []
    print(f'Transcribing audio from: {urls}')
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Start download and transcription operations and execute them concurrently
        future_to_url = {executor.submit(process_url, url): url for url in urls}
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                results.append(future.result())
            except Exception as exc:
                print(f'{url} generated an exception: {exc}')
    print(results)
    return results


   
def find_image_urls(data):
    image_urls = []  # Initialize an empty list to hold all image URLs
    for item in data:
        if "image_urls" in item:
            # Extend the list of image_urls with the URLs found in the current item
            image_urls.extend(item["image_urls"])
    # Return True if the list is not empty (indicating that at least one URL was found),
    # and the list of all found image URLs
    return bool(image_urls), image_urls
    

def text_to_speech(text, file_suffix=".mp3"):
    """
    Converts text to speech and saves the audio to a temporary file.

    Parameters:
    - text: The text to be converted to speech.
    - file_suffix: The suffix for the temporary file.

    Returns:  
    - The path to the saved speech file.    
    """
    with tempfile.NamedTemporaryFile(suffix=file_suffix, delete=False) as tmp_file:
        response = client.audio.speech.create(
            model="tts-1",
            voice="shimmer",
            input=text
        )
        response.stream_to_file(tmp_file.name)
        tmp_file_path = tmp_file.name

    return tmp_file_path

def smart_send_message(message, user_name):
    users = get_users()  # Assume this function returns a list of user dictionaries  
    matched_users = []
    similar_users = []

    for user in users:
        names = [user.get('real_name', ''), user.get('display_name', '')]
        for name in names:
            if user_name.lower() == name.lower():
                matched_users.append(user)
            elif user_name.lower() in name.lower().split():
                similar_users.append(user)

    if len(matched_users) == 1:
        user_id = matched_users[0]['user_id']
        send_slack_message(user_id, message, None)
        return f"Message sent to {matched_users[0]['real_name']} ({matched_users[0]['display_name']})."
    elif len(matched_users) > 1:
        return "Multiple matches found. Please confirm the user: " + ", ".join([f"{u['real_name']} ({u['display_name']})" for u in matched_users])
    elif similar_users:
        return "No exact match found. Did you mean: " + ", ".join([f"{u['real_name']} ({u['display_name']})" for u in similar_users]) + "?"
    else:
        return "No matching user found."


def get_channels(id = None):
    try:
        update_slack_conversations()
        # Perform a scan operation on the table to retrieve all channels  
        response = channels_table.scan()
        channels = response.get('Items', [])

        # Check if there are more channels to fetch
        while 'LastEvaluatedKey' in response:
            response = channels_table.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
            channels.extend(response.get('Items', []))

        return channels
    except Exception as e:
        print(f"Error fetching channels: {e}")
        return []
        

def update_slack_conversations():
    url = 'https://slack.com/api/conversations.list'
    headers = {
        'Authorization': f'Bearer {slack_bot_token}',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    params = {
        'types': 'public_channel,private_channel'
    }

    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        conversations_list = response.json().get('channels', [])
        for channel in conversations_list:
            channel_id = channel.get('id')
            try:
                # Retrieve the existing channel details from DynamoDB 
                existing_channel = channels_table.get_item(Key={'id': channel_id})
                
                if 'Item' in existing_channel:
                    print(f"Channel {channel_id} already exists in the database.")
                    # Existing channel: update if necessary
                    channels_table.update_item(
                        Key={'id': channel_id},
                        UpdateExpression="set #info.#name=:n, #info.#is_private=:p, #info.#num_members=:m",
                        ExpressionAttributeValues={
                            ':n': channel.get('name'),
                            ':p': channel.get('is_private'),
                            ':m': channel.get('num_members')
                        },
                        ExpressionAttributeNames={
                            "#info": "info",
                            "#name": "name",  # Using an expression attribute name as a placeholder
                            "#is_private": "is_private",
                            "#num_members": "num_members"
                        }
                    )
                else:
                    # New channel: add to the database      
                    channels_table.put_item(Item=channel)
                    print(f"Channel {channel_id} added to the database.")

            except ClientError as e:
                print(f"Error updating channel {channel_id}: {e}")
    else:
        print(f"Failed to retrieve conversations list: HTTP {response.status_code}")
        

def download_and_read_file(url, content_type):
    headers = {
        'Authorization': f'Bearer {slack_bot_token}',
        'Content-Type': 'image/x-www-form-urlencoded'
    }    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        # Extract the original file name and extension from the URL    
        parsed_url = urlparse(url)
        original_file_name = os.path.basename(unquote(parsed_url.path))
        _, file_extension = os.path.splitext(original_file_name)

        # Define the S3 bucket and folder where the file will be saved 
        bucket_name = docs_bucket_name
        folder_name = 'uploads'

        # Use the original file extension for the temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
            tmp_file.write(response.content)
            tmp_file.seek(0)
            #print(f"File Name: {tmp_file.name}")
            
            # Upload the file to S3 with the original file name    
            file_key = f"{folder_name}/{original_file_name}"
            s3_client = boto3.client('s3')
            s3_client.upload_file(tmp_file.name, bucket_name, file_key)

            if 'text/csv' in content_type:
                f = StringIO(response.text)
                reader = csv.reader(f)
                return '\n'.join([','.join(row) for row in reader])
            elif 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                workbook = openpyxl.load_workbook(tmp_file.name)
                sheet = workbook.active
                return '\n'.join([','.join([str(cell.value) for cell in row]) for row in sheet.rows])
            elif 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type:
                doc = Document(tmp_file.name)
                content = '\n'.join([p.text for p in doc.paragraphs])
                summary = rank_sentences(content, stopwords, max_sentences=50)
                return summary
            elif 'application/pdf' in content_type:
                pdf_reader = PyPDF2.PdfReader(tmp_file.name)
                text = ' '.join([page.extract_text() for page in pdf_reader.pages if page.extract_text()])
                summary = rank_sentences(text, stopwords, max_sentences=50)
                return summary
            elif 'text/plain' in content_type:
                with open(tmp_file.name, 'r') as f:
                    content = f.read()
                return content
            else:
                return 'Unsupported file type'
    except requests.exceptions.RequestException as e:
        return f'Error downloading file: {e}'
    except Exception as e:
        return f'Error processing file: {e}'
        

class MyFPDF(FPDF):
    def __init__(self, title):
        super().__init__()
        self.title = title
        self.inside_list = False

    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, self.title, 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def write_html(self, html):
        # Simple HTML parsing (extend this for more tags)
        html = html.replace('<strong>', '<b>').replace('</strong>', '</b>')
        html = html.replace('<em>', '<i>').replace('</em>', '</i>')

        # Split the HTML into parts
        parts = re.split('(<\/?[^>]+>)', html)
        for part in parts:
            # Detect and handle opening tags
            if part.startswith('<b>'):
                self.set_font('Arial', 'B', 11)
                continue
            if part.startswith('<i>'):
                self.set_font('Arial', 'I', 11)
                continue
            if part.startswith('<h1>'):
                self.set_font('Arial', 'B', 16)
                continue
            if part.startswith('<h2>'):
                self.set_font('Arial', 'B', 14)
                continue
            if part.startswith('<h3>'):
                self.set_font('Arial', 'B', 12)
                continue
            if part.startswith('<p>'):
                self.set_font('Arial', '', 11)
                continue
            if part.startswith('<ul>'):
                self.inside_list = True
                continue
            if part.startswith('<li>'):
                self.set_font('Arial', '', 11)
                part = '• ' + part[4:]
            
            # Detect and handle closing tags
            if part.startswith('</'):
                self.set_font('Arial', '', 11)
                if part.startswith('</h1>') or part.startswith('</h2>') or part.startswith('</h3>'):
                    self.ln(10)  # Add a line break after headers
                if part.startswith('</ul>'):
                    self.inside_list = False
                if part.startswith('</li>') and self.inside_list:
                    self.ln(5)  # Add a line break after list items
                continue

            # Write the actual text content   
            self.multi_cell(0, 8, part)


def replace_problematic_chars(text):
    # Replace common problematic Unicode characters with ASCII equivalents
    replacements = {
        '\u2018': "'",  # Left single quotation mark
        '\u2019': "'",  # Right single quotation mark
        '\u201c': '"',  # Left double quotation mark
        '\u201d': '"',  # Right double quotation mark
        '\u2013': '-',  # En dash
        '\u2014': '-',  # Em dash
        '\u2026': '...',  # Ellipsis
        '\u2022': '*',  # Bullet point
        '\u00A0': ' ',  # Non-breaking space
        '\u2010': '-',  # Hyphen
        '\u2012': '-',  # Figure dash
        '\u2015': '-',  # Horizontal bar
    }
    for original, replacement in replacements.items():
        text = text.replace(original, replacement)
    return text


def send_as_pdf(text, chat_id, title, ts=None):
    """
    Converts formatted text to PDF and uploads it to a Slack channel.  

    Parameters:
    - text: The text to be converted, may contain Markdown formatting.
    - chat_id: The ID of the Slack channel where the file will be uploaded.
    - title: The title of the file.
    - ts: The thread timestamp (optional).  
    """
    try:
        # Create a PDF object
        pdf = MyFPDF(title)
        pdf.add_page()
        pdf.set_font('Arial', '', 11)

        # Convert Markdown to HTML using markdown2
        text = text.replace("\\n\\n", "\\n")
        html_content = markdown2.markdown(text)

        # Replace problematic characters
        html_content = replace_problematic_chars(html_content)

        # Write HTML content to PDF
        pdf.write_html(html_content)

        # Save the PDF to a file in the /tmp directory
        pdf_path = f"/tmp/{title}.pdf"
        pdf.output(pdf_path)

        # Define the S3 bucket and folder where the file will be saved
        bucket_name = docs_bucket_name
        folder_name = 'uploads'

        # Upload the file to S3 with the original file name  
        file_key = f"{folder_name}/{title}.pdf"
        s3_client = boto3.client('s3')
        s3_client.upload_file(pdf_path, bucket_name, file_key)

        # Upload the PDF to Slack
        send_file_to_slack(pdf_path, chat_id, title, ts)

    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        # Clean up the temporary PDF file 
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
    

def list_files(folder_prefix='uploads'):
    """
    Lists the files in a specified folder in the 'docs_bucket_name' S3 bucket.

    Parameters:
    - folder_prefix: The prefix of the folder whose files you want to list (optional). This should end with a slash ('/') if provided.   

    Returns:
    - A dictionary where each key is the file name and the value is the object URL.
    """
    s3 = boto3.client('s3')
    response = s3.list_objects_v2(Bucket=docs_bucket_name, Prefix=folder_prefix)

    files = {}
    if 'Contents' in response:
        for obj in response['Contents']:
            if not obj['Key'].endswith('/'):  # Exclude any subfolders
                file_url = f"https://{docs_bucket_name}.s3.amazonaws.com/{obj['Key']}"
                file_name = obj['Key'].split('/')[-1]  # Extract the file name from the key
                files[file_name] = file_url

    return files
    

def ask_bighead(prompt, model_option='pro'):
    model_map = {
        'ultra': 'gemini-1.0-ultra-latest',
        'pro': 'gemini-1.5-pro-latest',
        'vision': 'gemini-pro-vision'
    }
    model = model_map.get(model_option.lower())
    if not model:
        raise ValueError(f'Invalid model option: {model_option}. Choose from "pro", or "vision".')

    url = f'https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={gemini_api_key}'
    headers = {'Content-Type': 'application/json'}
    data = {
        'contents': [{
            'parts': [{
                'text': prompt
            }]
        }]
    }
    try:
        response = requests.post(url, headers=headers, json=data)
        response.raise_for_status()  # Raises an HTTPError if the status is 4xx or 5xx   
        response_json = response.json()
        print(json.dumps(response_json))
        if 'candidates' in response_json and response_json['candidates']:
            first_candidate = response_json['candidates'][0]
            if 'content' in first_candidate and 'parts' in first_candidate['content'] and first_candidate['content']['parts']:
                response_text = first_candidate['content']['parts'][0]['text']
                #print(response_text)
                return to_markdown(response_text)
        return None
    except requests.exceptions.HTTPError as e:
        return f'HTTP error occurred: {e.response.text}'
    except requests.exceptions.RequestException as e:
        return f'Request error occurred: {e}'
    except Exception as e:
        return f'An error occurred: {e}'
        
        
def to_markdown(text):
    text = text.replace('•', '  *')
    indented_text = textwrap.indent(text, '> ', predicate=lambda _: True)
    html = markdown2.markdown(indented_text)
    return html


def manage_mute_status(status=None):
    table = channels_table
    
    if status is not None:
        # Initialize status_bool based on the type and value of status
        if isinstance(status, bool):
            status_bool = status
        elif isinstance(status, str):
            status = status.strip()  # Remove any leading/trailing whitespace
            if status.lower() in ['true', 'false']:
                status_bool = status.lower() == 'true'
            else:
                raise ValueError("String status must be 'true' or 'false' (case insensitive).")
        else:
            raise TypeError("Status must be provided as either a boolean or a string.")

        try:
            response = table.update_item(
                Key={
                    'id': chat_id
                },
                UpdateExpression='SET maria_status = :val',
                ExpressionAttributeValues={
                    ':val': status_bool
                },
                ReturnValues='UPDATED_NEW'
            )
            current_status = "true" if status_bool else "false"
            return [status_bool, f"Current mute status: {current_status}"]
        except ClientError as e:
            print(f"An error occurred: {e}")
            raise
    else:
        try:
            response = table.get_item(
                Key={
                    'id': chat_id
                }
            )
            item = response.get('Item', {})
            maria_status = item.get('maria_status', None)

            if maria_status is None:
                status_bool = False  # Assuming False as default if status doesn't exist
                current_status = "false"
                print("The 'maria_status' attribute does not exist for this record.")
            else:
                status_bool = maria_status  # Assuming maria_status is stored as a boolean
                current_status = "true" if maria_status else "false"
            
            return [status_bool, f"Current mute status: {current_status}"]
        except ClientError as e:
            print(f"An error occurred: {e}")
            raise
        
        
def solve_maths(code: str, **params) -> dict:
    """
    Execute the given code and return the result.

    Parameters:
    - code (str): The Python code to execute.
    - **params: Any parameters that the code might need.

    Returns:
    - dict: A dictionary containing the result or an error message.
    """
    exec_env = {}
    exec_env.update(params)
    
    try:
        exec(code, {}, exec_env)
        
        # Filter out non-serializable objects
        serializable_result = {key: value for key, value in exec_env.items() if is_serializable(value)}
        
        return {"status": "success", "result": serializable_result}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def is_serializable(value):
    """Helper function to check if a value is serializable."""
    try:
        json.dumps(value)
        return True
    except (TypeError, OverflowError):
        return False


def clean_website_data(raw_text):
    """
    Cleans up raw website text data, removing common HTML artifacts and excess whitespace.
    """
    try:
        # Remove HTML tags (basic HTML tag removal)
        cleaned_text = re.sub('<[^<]+?>', '', raw_text)

        # Remove multiple spaces and newlines, and then trim
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        
        #Remove non-printing characters
        cleaned_text = ''.join(c for c in cleaned_text if c.isprintable())
        return cleaned_text

    except Exception as e:
        return json.dumps({"error": f"Error processing text: {str(e)}"})


def convert_to_wav_in_memory(m4a_data):
    """
    Converts in-memory M4A/MP4 audio data to WAV format.
    Assumes the input is raw PCM audio data.

    Args:
        m4a_data (bytes): Audio data from M4A/MP4.

    Returns:
        bytes: Converted WAV data in memory.

    Raises:
        ValueError: If the input audio data is invalid.
        RuntimeError: If WAV conversion fails.
    """
    # Create an in-memory buffer for the output WAV file
    wav_buffer = BytesIO()

    try:
        # Open the buffer as a WAV file for writing
        with wave.open(wav_buffer, "wb") as wav_file:
            # Set the required WAV parameters
            wav_file.setnchannels(2)  # Stereo
            wav_file.setsampwidth(2)  # 16-bit samples (2 bytes per sample)
            wav_file.setframerate(44100)  # 44.1 kHz sample rate

            # Write the raw PCM audio data to the WAV file
            wav_file.writeframes(m4a_data)

        # Return WAV data as bytes
        wav_buffer.seek(0)  # Rewind buffer to the beginning
        return wav_buffer.read()

    except wave.Error as e:
        raise ValueError("Invalid WAV data or parameters.") from e

    except Exception as e:
        raise RuntimeError("An error occurred during WAV conversion.") from e

    finally:
        wav_buffer.close()


def normalize_message(response):
    # If it's a dictionary (raw JSON format)
    if isinstance(response, dict):
        message = response['choices'][0]['message']
        return type('Message', (), {
            'content': message.get('content'),
            'audio': message.get('audio'),
            'tool_calls': message.get('tool_calls')
        })
    # If it's already an OpenAI object
    return response.choices[0].message
