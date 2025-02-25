import os
import json
import boto3
import mimetypes
import tempfile
import datetime
import requests
import markdown2
from typing import Dict, Any, Optional
from urllib.parse import urlparse, unquote
from io import BytesIO
from fpdf import FPDF
from docx import Document
import PyPDF2
import csv
import openpyxl

from ..core.config import (
    SLACK_BOT_TOKEN,
    DOCS_BUCKET_NAME,
    IMAGE_BUCKET_NAME
)
from ..core.error_handlers import FileOperationError

class MyFPDF(FPDF):
    """Extended FPDF class with custom header and footer"""
    def __init__(self, title):
        super().__init__()
        self.title = title
        self.inside_list = False

    def header(self):
        """Add custom header to each page"""
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, self.title, 0, 1, 'C')

    def footer(self):
        """Add custom footer to each page"""
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def write_html(self, html):
        """Write HTML content to PDF with proper formatting"""
        # Simple HTML parsing (extend this for more tags)
        html = html.replace('<strong>', '<b>').replace('</strong>', '</b>')
        html = html.replace('<em>', '<i>').replace('</em>', '</i>')

        # Split the HTML into parts
        parts = re.split('(<\/?[^>]+>)', html)
        for part in parts:
            # Detect and handle opening tags
            if part.startswith('<b>'):
                self.set_font('Arial', 'B', 11)
                continue
            if part.startswith('<i>'):
                self.set_font('Arial', 'I', 11)
                continue
            if part.startswith('<h1>'):
                self.set_font('Arial', 'B', 16)
                continue
            if part.startswith('<h2>'):
                self.set_font('Arial', 'B', 14)
                continue
            if part.startswith('<h3>'):
                self.set_font('Arial', 'B', 12)
                continue
            if part.startswith('<p>'):
                self.set_font('Arial', '', 11)
                continue
            if part.startswith('<ul>'):
                self.inside_list = True
                continue
            if part.startswith('<li>'):
                self.set_font('Arial', '', 11)
                part = '• ' + part[4:]
            
            # Detect and handle closing tags
            if part.startswith('</'):
                self.set_font('Arial', '', 11)
                if part.startswith('</h1>') or part.startswith('</h2>') or part.startswith('</h3>'):
                    self.ln(10)  # Add a line break after headers
                if part.startswith('</ul>'):
                    self.inside_list = False
                if part.startswith('</li>') and self.inside_list:
                    self.ln(5)  # Add a line break after list items
                continue

            # Write the actual text content
            self.multi_cell(0, 8, part)

def download_and_read_file(url: str, content_type: str) -> str:
    """
    Download and read file content based on its type.

    Args:
        url: File URL
        content_type: MIME type of the file

    Returns:
        String containing file content or summary

    Raises:
        FileOperationError: If file operations fail
    """
    headers = {
        'Authorization': f'Bearer {SLACK_BOT_TOKEN}',
        'Content-Type': 'image/x-www-form-urlencoded'
    }
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        # Extract the original file name and extension
        parsed_url = urlparse(url)
        original_file_name = os.path.basename(unquote(parsed_url.path))
        _, file_extension = os.path.splitext(original_file_name)

        # Create temporary file with original extension
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
            tmp_file.write(response.content)
            tmp_file.seek(0)

            # Upload to S3
            file_key = f"uploads/{original_file_name}"
            s3_client = boto3.client('s3')
            s3_client.upload_file(tmp_file.name, DOCS_BUCKET_NAME, file_key)

            # Process file based on content type
            if 'text/csv' in content_type:
                f = BytesIO(response.content)
                reader = csv.reader(f.read().decode().splitlines())
                return '\n'.join([','.join(row) for row in reader])

            elif 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                workbook = openpyxl.load_workbook(tmp_file.name)
                sheet = workbook.active
                return '\n'.join([','.join([str(cell.value) for cell in row]) for row in sheet.rows])

            elif 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type:
                doc = Document(tmp_file.name)
                content = '\n'.join([p.text for p in doc.paragraphs])
                summary = rank_sentences(content, stopwords, max_sentences=50)  # This needs to be imported
                return summary

            elif 'application/pdf' in content_type:
                pdf_reader = PyPDF2.PdfReader(tmp_file.name)
                text = ' '.join([page.extract_text() for page in pdf_reader.pages if page.extract_text()])
                summary = rank_sentences(text, stopwords, max_sentences=50)  # This needs to be imported
                return summary

            elif 'text/plain' in content_type:
                with open(tmp_file.name, 'r') as f:
                    return f.read()

            else:
                return 'Unsupported file type'

    except Exception as e:
        raise FileOperationError(
            message="Failed to process file",
            status_code=500,
            details={"error": str(e)}
        )
    finally:
        if 'tmp_file' in locals():
            os.unlink(tmp_file.name)

def upload_document_to_s3(
    document_content: bytes,
    content_type: str,
    document_url: str
) -> str:
    """
    Upload a document to S3.

    Args:
        document_content: Raw document content
        content_type: MIME type
        document_url: Original document URL

    Returns:
        S3 URL of uploaded document

    Raises:
        FileOperationError: If upload fails
    """
    try:
        document_extension = mimetypes.guess_extension(content_type) or '.bin'
        s3_object_name = f"Document_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}{document_extension}"

        s3_client = boto3.client('s3')
        s3_client.put_object(
            Body=document_content,
            Bucket=DOCS_BUCKET_NAME,
            Key=s3_object_name
        )

        s3_url = f"https://{DOCS_BUCKET_NAME}.s3.amazonaws.com/{s3_object_name}"
        return s3_url
    except Exception as e:
        raise FileOperationError(
            message="Failed to upload document to S3",
            status_code=500,
            details={"error": str(e)}
        )

def upload_image_to_s3(image_url: str, bucket_name: str) -> str:
    """
    Download image from URL and upload to S3.

    Args:
        image_url: URL of the image
        bucket_name: S3 bucket name

    Returns:
        S3 URL of uploaded image

    Raises:
        FileOperationError: If download or upload fails
    """
    try:
        headers = {
            'Authorization': f'Bearer {SLACK_BOT_TOKEN}',
            'Content-Type': 'image/x-www-form-urlencoded'
        }
        response = requests.get(image_url, headers=headers)
        response.raise_for_status()

        image_content = BytesIO(response.content)
        file_extension = os.path.splitext(image_url)[1]
        s3_object_name = f"Image_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}{file_extension}"

        s3_client = boto3.client('s3')
        s3_client.upload_fileobj(image_content, bucket_name, s3_object_name)

        s3_url = f"https://{bucket_name}.s3.amazonaws.com/{s3_object_name}"
        return s3_url
    except Exception as e:
        raise FileOperationError(
            message="Failed to upload image to S3",
            status_code=500,
            details={"error": str(e)}
        )

def list_files(folder_prefix: str = 'uploads') -> Dict[str, str]:
    """
    List files in S3 bucket folder.

    Args:
        folder_prefix: Folder path in S3 bucket

    Returns:
        Dictionary mapping file names to URLs

    Raises:
        FileOperationError: If listing fails
    """
    try:
        s3_client = boto3.client('s3')
        response = s3_client.list_objects_v2(
            Bucket=DOCS_BUCKET_NAME,
            Prefix=folder_prefix
        )

        files = {}
        if 'Contents' in response:
            for obj in response['Contents']:
                if not obj['Key'].endswith('/'):  # Exclude directories
                    file_url = f"https://{DOCS_BUCKET_NAME}.s3.amazonaws.com/{obj['Key']}"
                    file_name = obj['Key'].split('/')[-1]
                    files[file_name] = file_url

        return files
    except Exception as e:
        raise FileOperationError(
            message="Failed to list files",
            status_code=500,
            details={"error": str(e)}
        )

def send_as_pdf(text: str, chat_id: str, title: str, ts: Optional[str] = None) -> None:
    """
    Convert text to PDF and send to Slack channel.

    Args:
        text: Text content to convert
        chat_id: Slack channel ID
        title: PDF title
        ts: Thread timestamp (optional)

    Raises:
        FileOperationError: If PDF creation or upload fails
    """
    try:
        # Create a PDF object
        pdf = MyFPDF(title)
        pdf.add_page()
        pdf.set_font('Arial', '', 11)

        # Convert Markdown to HTML
        text = text.replace("\\n\\n", "\\n")
        html_content = markdown2.markdown(text)

        # Replace problematic characters
        html_content = replace_problematic_chars(html_content)  # This needs to be imported

        # Write HTML content to PDF
        pdf.write_html(html_content)

        # Save the PDF to a temporary file
        pdf_path = f"/tmp/{title}.pdf"
        pdf.output(pdf_path)

        try:
            # Upload the PDF to S3
            file_key = f"uploads/{title}.pdf"
            s3_client = boto3.client('s3')
            s3_client.upload_file(pdf_path, DOCS_BUCKET_NAME, file_key)

            # Send the PDF to Slack
            send_file_to_slack(pdf_path, chat_id, title, ts)  # This needs to be imported

        finally:
            # Clean up the temporary PDF file
            if os.path.exists(pdf_path):
                os.remove(pdf_path)

    except Exception as e:
        raise FileOperationError(
            message="Failed to create or send PDF",
            status_code=500,
            details={"error": str(e)}
        )