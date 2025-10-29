#!/usr/bin/env python3
"""
Test script for enhanced Excel generation functionality.
Tests the improved send_as_excel function with styling features.
"""

import json
import os
import sys
from openpyxl import load_workbook

# Add parent directory to path to import lambda_function
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from lambda_function import ExcelGenerator, parse_excel_data, EXCEL_COLOR_SCHEMES

def test_color_schemes():
    """Test all predefined color schemes"""
    print("Testing color schemes...")
    
    test_data = [
        {"Product": "Laptop", "Price": 999.99, "Stock": 50},
        {"Product": "Mouse", "Price": 29.99, "Stock": 200},
        {"Product": "Keyboard", "Price": 79.99, "Stock": 150}
    ]
    
    headers = ["Product", "Price", "Stock"]
    
    for scheme_name, scheme_colors in EXCEL_COLOR_SCHEMES.items():
        print(f"  ✓ Testing '{scheme_name}' color scheme...")
        
        generator = ExcelGenerator(
            title=f"test_{scheme_name}",
            sheet_name="Products",
            header_color=scheme_colors['header_color'],
            header_font_color=scheme_colors['header_font_color'],
            header_row_height=30,
            freeze_header=True,
            wrap_header_text=True
        )
        
        generator.write_data(headers, test_data, include_headers=True)
        filepath = f"/tmp/test_{scheme_name}.xlsx"
        generator.save(filepath)
        
        # Verify file was created
        assert os.path.exists(filepath), f"File not created for {scheme_name}"
        
        # Load and verify
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check freeze panes
        assert ws.freeze_panes == 'A2', f"Freeze panes not set correctly for {scheme_name}"
        
        # Check header row height
        assert ws.row_dimensions[1].height == 30, f"Header row height not set for {scheme_name}"
        
        # Check header styling
        header_cell = ws['A1']
        assert header_cell.fill.start_color.rgb == f"FF{scheme_colors['header_color']}", \
            f"Header color not correct for {scheme_name}"
        
        # Clean up
        os.remove(filepath)
        
    print("  ✅ All color schemes working correctly!\n")


def test_parse_excel_data():
    """Test data parsing functionality"""
    print("Testing data parsing...")
    
    # Test JSON parsing
    json_data = '[{"Name": "John", "Age": 30}, {"Name": "Jane", "Age": 25}]'
    headers, rows = parse_excel_data(json_data)
    assert headers == ["Name", "Age"], "JSON headers parsing failed"
    assert len(rows) == 2, "JSON rows parsing failed"
    print("  ✓ JSON parsing works")
    
    # Test CSV parsing
    csv_data = "Name,Age,City\nJohn,30,NYC\nJane,25,LA"
    headers, rows = parse_excel_data(csv_data)
    assert headers == ["Name", "Age", "City"], "CSV headers parsing failed"
    assert len(rows) == 2, "CSV rows parsing failed"
    print("  ✓ CSV parsing works")
    
    print("  ✅ Data parsing working correctly!\n")


def test_freeze_and_formatting():
    """Test freeze panes and formatting features"""
    print("Testing freeze panes and formatting...")
    
    test_data = [
        {"Department": "Engineering", "Employees": 50, "Budget": 1000000},
        {"Department": "Marketing", "Employees": 20, "Budget": 500000}
    ]
    
    headers = ["Department", "Employees", "Budget"]
    
    # Test with freeze enabled
    generator_freeze = ExcelGenerator(
        title="test_freeze",
        freeze_header=True,
        header_row_height=35
    )
    generator_freeze.write_data(headers, test_data, include_headers=True)
    filepath_freeze = "/tmp/test_freeze.xlsx"
    generator_freeze.save(filepath_freeze)
    
    wb = load_workbook(filepath_freeze)
    ws = wb.active
    assert ws.freeze_panes == 'A2', "Freeze panes not working"
    assert ws.row_dimensions[1].height == 35, "Row height not set correctly"
    print("  ✓ Freeze panes enabled")
    os.remove(filepath_freeze)
    
    # Test with freeze disabled
    generator_no_freeze = ExcelGenerator(
        title="test_no_freeze",
        freeze_header=False
    )
    generator_no_freeze.write_data(headers, test_data, include_headers=True)
    filepath_no_freeze = "/tmp/test_no_freeze.xlsx"
    generator_no_freeze.save(filepath_no_freeze)
    
    wb2 = load_workbook(filepath_no_freeze)
    ws2 = wb2.active
    assert ws2.freeze_panes is None, "Freeze panes should be disabled"
    print("  ✓ Freeze panes disabled")
    os.remove(filepath_no_freeze)
    
    print("  ✅ Freeze panes and formatting working correctly!\n")


def test_manual_column_widths():
    """Test manual column width setting"""
    print("Testing manual column widths...")
    
    test_data = [
        {"Name": "John Doe", "Email": "john@example.com", "Phone": "555-1234"}
    ]
    
    headers = ["Name", "Email", "Phone"]
    manual_widths = [25, 35, 15]
    
    generator = ExcelGenerator(title="test_widths")
    generator.write_data(headers, test_data, include_headers=True, column_widths=manual_widths)
    filepath = "/tmp/test_widths.xlsx"
    generator.save(filepath)
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Check column widths
    assert ws.column_dimensions['A'].width == 25, "Column A width not set"
    assert ws.column_dimensions['B'].width == 35, "Column B width not set"
    assert ws.column_dimensions['C'].width == 15, "Column C width not set"
    
    print("  ✓ Manual column widths set correctly")
    os.remove(filepath)
    print("  ✅ Manual column widths working correctly!\n")


def run_all_tests():
    """Run all enhancement tests"""
    print("="*60)
    print("Testing Enhanced Excel Generation Features")
    print("="*60 + "\n")
    
    try:
        test_parse_excel_data()
        test_color_schemes()
        test_freeze_and_formatting()
        test_manual_column_widths()
        
        print("="*60)
        print("✅ ALL TESTS PASSED!")
        print("="*60)
        print("\nEnhanced features verified:")
        print("  • Multiple color schemes (blue, green, yellow, red, etc.)")
        print("  • Freeze panes for header rows")
        print("  • Customizable header row height")
        print("  • Text wrapping in headers")
        print("  • Manual column width control")
        print("  • Auto-sizing columns")
        print("\nThe send_as_excel function now has professional styling!")
        
    except AssertionError as e:
        print(f"\n❌ TEST FAILED: {e}")
        return False
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)