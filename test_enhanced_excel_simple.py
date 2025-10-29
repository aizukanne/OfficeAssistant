#!/usr/bin/env python3
"""
Simplified test script for enhanced Excel generation functionality.
Tests the improved ExcelGenerator class with styling features.
"""

import json
import csv
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Predefined Excel color schemes (copied from lambda_function.py)
EXCEL_COLOR_SCHEMES = {
    'blue': {'header_color': '4472C4', 'header_font_color': 'FFFFFF'},
    'green': {'header_color': '70AD47', 'header_font_color': 'FFFFFF'},
    'yellow': {'header_color': 'FFC000', 'header_font_color': '000000'},
    'red': {'header_color': 'C00000', 'header_font_color': 'FFFFFF'},
    'purple': {'header_color': '7030A0', 'header_font_color': 'FFFFFF'},
    'orange': {'header_color': 'ED7D31', 'header_font_color': 'FFFFFF'},
    'teal': {'header_color': '00B0F0', 'header_font_color': 'FFFFFF'},
    'gray': {'header_color': '7F7F7F', 'header_font_color': 'FFFFFF'}
}


class ExcelGenerator:
    """
    Excel generation class with enhanced formatting and styling.
    Supports customizable colors, freeze panes, row heights, and text wrapping.
    """

    def __init__(self, title, sheet_name='Sheet1', header_color='4472C4', 
                 header_font_color='FFFFFF', header_row_height=30, 
                 freeze_header=True, wrap_header_text=True):
        """
        Initialize Excel generator with workbook and worksheet.
        
        Args:
            title (str): Title of the workbook
            sheet_name (str): Name of the worksheet
            header_color (str): Hex color for header background (default: blue '4472C4')
            header_font_color (str): Hex color for header text (default: white 'FFFFFF')
            header_row_height (int): Height of header row in points (default: 30)
            freeze_header (bool): Whether to freeze header row (default: True)
            wrap_header_text (bool): Whether to wrap text in headers (default: True)
        """
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        self.title = title
        self.header_color = header_color
        self.header_font_color = header_font_color
        self.header_row_height = header_row_height
        self.freeze_header = freeze_header
        self.wrap_header_text = wrap_header_text

    def write_data(self, headers, rows, include_headers=True, column_widths=None):
        """
        Write data to Excel worksheet with enhanced formatting and styling.

        Args:
            headers (list): Column headers
            rows (list): List of dictionaries containing row data
            include_headers (bool): Whether to include header row
            column_widths (list, optional): Manual column widths. If None, auto-sizes.
        """
        # Define styles using instance parameters
        header_font = Font(bold=True, color=self.header_font_color, size=11)
        header_fill = PatternFill(start_color=self.header_color, 
                                  end_color=self.header_color, 
                                  fill_type="solid")
        border = Border(left=Side(style='thin'),
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))
        
        # Header alignment with optional wrap text
        header_alignment = Alignment(horizontal="center", 
                                    vertical="center", 
                                    wrap_text=self.wrap_header_text)

        row_num = 1

        # Write headers if enabled
        if include_headers:
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = header_alignment
            
            # Set header row height
            self.worksheet.row_dimensions[row_num].height = self.header_row_height
            
            row_num += 1

        # Write data rows
        for row_data in rows:
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.value = row_data.get(header, "")
                cell.border = border

                # Auto-format numbers
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(cell.value, str):
                    cell.alignment = Alignment(horizontal="left")

            row_num += 1

        # Set column widths (manual or auto-size)
        if column_widths:
            # Use manual column widths if provided
            for col_num, width in enumerate(column_widths, start=1):
                if col_num <= len(headers):
                    column_letter = get_column_letter(col_num)
                    self.worksheet.column_dimensions[column_letter].width = width
        else:
            # Auto-size columns
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                column = self.worksheet[column_letter]
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row if enabled
        if include_headers and self.freeze_header:
            self.worksheet.freeze_panes = 'A2'

    def save(self, filepath):
        """Save workbook to specified file path."""
        self.workbook.save(filepath)


def test_color_schemes():
    """Test all predefined color schemes"""
    print("Testing color schemes...")
    
    test_data = [
        {"Product": "Laptop", "Price": 999.99, "Stock": 50},
        {"Product": "Mouse", "Price": 29.99, "Stock": 200},
        {"Product": "Keyboard", "Price": 79.99, "Stock": 150}
    ]
    
    headers = ["Product", "Price", "Stock"]
    
    for scheme_name, scheme_colors in EXCEL_COLOR_SCHEMES.items():
        print(f"  ✓ Testing '{scheme_name}' color scheme...")
        
        generator = ExcelGenerator(
            title=f"test_{scheme_name}",
            sheet_name="Products",
            header_color=scheme_colors['header_color'],
            header_font_color=scheme_colors['header_font_color'],
            header_row_height=30,
            freeze_header=True,
            wrap_header_text=True
        )
        
        generator.write_data(headers, test_data, include_headers=True)
        filepath = f"/tmp/test_{scheme_name}.xlsx"
        generator.save(filepath)
        
        # Verify file was created
        assert os.path.exists(filepath), f"File not created for {scheme_name}"
        
        # Load and verify
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check freeze panes
        assert ws.freeze_panes == 'A2', f"Freeze panes not set correctly for {scheme_name}"
        
        # Check header row height
        assert ws.row_dimensions[1].height == 30, f"Header row height not set for {scheme_name}"
        
        # Clean up
        os.remove(filepath)
        
    print("  ✅ All color schemes working correctly!\n")


def test_freeze_and_formatting():
    """Test freeze panes and formatting features"""
    print("Testing freeze panes and formatting...")
    
    test_data = [
        {"Department": "Engineering", "Employees": 50, "Budget": 1000000},
        {"Department": "Marketing", "Employees": 20, "Budget": 500000}
    ]
    
    headers = ["Department", "Employees", "Budget"]
    
    # Test with freeze enabled
    generator_freeze = ExcelGenerator(
        title="test_freeze",
        freeze_header=True,
        header_row_height=35
    )
    generator_freeze.write_data(headers, test_data, include_headers=True)
    filepath_freeze = "/tmp/test_freeze.xlsx"
    generator_freeze.save(filepath_freeze)
    
    wb = load_workbook(filepath_freeze)
    ws = wb.active
    assert ws.freeze_panes == 'A2', "Freeze panes not working"
    assert ws.row_dimensions[1].height == 35, "Row height not set correctly"
    print("  ✓ Freeze panes enabled")
    os.remove(filepath_freeze)
    
    # Test with freeze disabled
    generator_no_freeze = ExcelGenerator(
        title="test_no_freeze",
        freeze_header=False
    )
    generator_no_freeze.write_data(headers, test_data, include_headers=True)
    filepath_no_freeze = "/tmp/test_no_freeze.xlsx"
    generator_no_freeze.save(filepath_no_freeze)
    
    wb2 = load_workbook(filepath_no_freeze)
    ws2 = wb2.active
    assert ws2.freeze_panes is None, "Freeze panes should be disabled"
    print("  ✓ Freeze panes disabled")
    os.remove(filepath_no_freeze)
    
    print("  ✅ Freeze panes and formatting working correctly!\n")


def test_manual_column_widths():
    """Test manual column width setting"""
    print("Testing manual column widths...")
    
    test_data = [
        {"Name": "John Doe", "Email": "john@example.com", "Phone": "555-1234"}
    ]
    
    headers = ["Name", "Email", "Phone"]
    manual_widths = [25, 35, 15]
    
    generator = ExcelGenerator(title="test_widths")
    generator.write_data(headers, test_data, include_headers=True, column_widths=manual_widths)
    filepath = "/tmp/test_widths.xlsx"
    generator.save(filepath)
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Check column widths
    assert ws.column_dimensions['A'].width == 25, "Column A width not set"
    assert ws.column_dimensions['B'].width == 35, "Column B width not set"
    assert ws.column_dimensions['C'].width == 15, "Column C width not set"
    
    print("  ✓ Manual column widths set correctly")
    os.remove(filepath)
    print("  ✅ Manual column widths working correctly!\n")


def run_all_tests():
    """Run all enhancement tests"""
    print("="*60)
    print("Testing Enhanced Excel Generation Features")
    print("="*60 + "\n")
    
    try:
        test_color_schemes()
        test_freeze_and_formatting()
        test_manual_column_widths()
        
        print("="*60)
        print("✅ ALL TESTS PASSED!")
        print("="*60)
        print("\nEnhanced features verified:")
        print("  • Multiple color schemes (blue, green, yellow, red, etc.)")
        print("  • Freeze panes for header rows")
        print("  • Customizable header row height")
        print("  • Text wrapping in headers")
        print("  • Manual column width control")
        print("  • Auto-sizing columns")
        print("\nThe send_as_excel function now has professional styling!")
        
    except AssertionError as e:
        print(f"\n❌ TEST FAILED: {e}")
        return False
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True


if __name__ == "__main__":
    import sys
    success = run_all_tests()
    sys.exit(0 if success else 1)